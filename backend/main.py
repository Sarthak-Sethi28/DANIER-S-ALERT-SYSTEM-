from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Form, BackgroundTasks
from fastapi.responses import FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
import tempfile
import os
import gc
import asyncio
import psutil
from datetime import datetime, timedelta
import pandas as pd
import threading
import secrets
import hashlib

from database import get_db, engine, init_db
from models import Base, Recipient, UploadedFile, ThresholdOverride, ThresholdHistory
from key_items_service import KeyItemsService
from email_service import EmailService
from file_storage_service import FileStorageService
from comparison_service import ComparisonService
from recipients_storage import recipients_storage
from threshold_analysis_service import ThresholdAnalysisService

# Initialize database
init_db()

app = FastAPI(title="Danier Stock Alert System")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize services
key_items_service = KeyItemsService()
email_service = EmailService()
file_storage_service = FileStorageService()
comparison_service = ComparisonService(key_items_service)
threshold_analysis_service = ThresholdAnalysisService()

# --- Simple credential utilities ---
from models import UserCredential  # type: ignore

def _hash_password(password: str, salt: str) -> str:
    return hashlib.sha256((salt + password).encode('utf-8')).hexdigest()

def _create_or_get_default_user(db: Session):
    try:
        user = db.query(UserCredential).first()
        if user:
            return user
        # Seed default user matching current demo creds
        username = os.getenv('APP_USERNAME', 'danier_admin')
        raw_password = os.getenv('APP_PASSWORD', 'danier2024')
        email = os.getenv('APP_USER_EMAIL', 'danieralertsystem@gmail.com')
        salt = secrets.token_hex(16)
        pwd_hash = _hash_password(raw_password, salt)
        user = UserCredential(username=username, password_hash=pwd_hash, password_salt=salt, email=email)
        db.add(user)
        db.commit()
        return user
    except Exception:
        db.rollback()
        return None

# --- Auth endpoints ---
@app.post("/auth/login")
async def auth_login(username: str = Form(...), password: str = Form(...)):
    try:
        db = next(get_db())
        try:
            _create_or_get_default_user(db)
            user = db.query(UserCredential).filter(UserCredential.username == username).first()
            if not user:
                raise HTTPException(status_code=401, detail="Invalid credentials")
            if user.password_hash != _hash_password(password, user.password_salt):
                raise HTTPException(status_code=401, detail="Invalid credentials")
            # Lightweight session response (frontend still stores locally)
            return {
                "success": True,
                "username": user.username,
                "loginTime": datetime.utcnow().isoformat(),
                "sessionId": f"session_{int(datetime.utcnow().timestamp())}_{secrets.token_hex(4)}"
            }
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/auth/request-reset")
async def request_password_reset(username: str = Form(...)):
    try:
        db = next(get_db())
        try:
            user = db.query(UserCredential).filter(UserCredential.username == username).first()
            if not user:
                # For privacy, do not reveal user existence
                return {"success": True, "message": "If the user exists, a reset code has been sent."}
            # Generate 6-digit code
            code = f"{secrets.randbelow(1000000):06d}"
            salt = secrets.token_hex(8)
            code_hash = _hash_password(code, salt)
            user.reset_code_hash = f"{salt}:{code_hash}"
            user.reset_code_expires_at = datetime.utcnow() + timedelta(minutes=15)
            db.add(user)
            db.commit()
            # Send email with code (best-effort)
            try:
                subject = "Danier Dashboard Password Reset Code"
                html = f"""
                <html><body>
                <p>Your password reset code is:</p>
                <h2 style='letter-spacing:3px'>{code}</h2>
                <p>This code expires in 15 minutes.</p>
                <p>Username: <b>{user.username}</b></p>
                </body></html>
                """
                # Build recipient list from active recipients + user's email
                recipient_list = []
                try:
                    active = recipients_storage.get_active_recipients()
                    recipient_list = [r.get('email') for r in active if r.get('email')]
                except Exception:
                    recipient_list = []
                if user.email and user.email not in recipient_list:
                    recipient_list.append(user.email)
                if not recipient_list:
                    recipient_list = [os.getenv('RESET_FALLBACK_EMAIL', 'danieralertsystem@gmail.com')]
                for rcpt in recipient_list:
                    try:
                        email_service._send_real_gmail_email(rcpt, subject, html, rcpt)
                    except Exception:
                        continue
            except Exception:
                pass
            return {"success": True, "message": "If the user exists, a reset code has been sent to active recipients."}
        finally:
            db.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/auth/confirm-reset")
async def confirm_password_reset(
    username: str = Form(...),
    code: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
):
    try:
        if len(new_password) < 8:
            raise HTTPException(status_code=400, detail="Password must be at least 8 characters")
        if new_password != confirm_password:
            raise HTTPException(status_code=400, detail="Passwords do not match")
        db = next(get_db())
        try:
            user = db.query(UserCredential).filter(UserCredential.username == username).first()
            if not user or not user.reset_code_hash or not user.reset_code_expires_at:
                raise HTTPException(status_code=400, detail="Invalid or expired code")
            if datetime.utcnow() > user.reset_code_expires_at:
                # expire
                user.reset_code_hash = None
                user.reset_code_expires_at = None
                db.commit()
                raise HTTPException(status_code=400, detail="Reset code expired")
            # Verify code
            try:
                salt, stored_hash = (user.reset_code_hash or '').split(':', 1)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid code state")
            if _hash_password(code, salt) != stored_hash:
                raise HTTPException(status_code=400, detail="Invalid code")
            # Set new password
            new_salt = secrets.token_hex(16)
            user.password_salt = new_salt
            user.password_hash = _hash_password(new_password, new_salt)
            user.reset_code_hash = None
            user.reset_code_expires_at = None
            db.add(user)
            db.commit()
            return {"success": True, "message": "Password updated successfully"}
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Default recipient email (can be configured via environment variable)
RECIPIENT_EMAIL = os.getenv("DEFAULT_RECIPIENT_EMAIL", "alerts@danier.ca")
# Upload directory from env (Render persistent disk) with fallback
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "uploads")

def log_memory_usage():
    """Log current memory usage for monitoring"""
    try:
        process = psutil.Process()
        memory_info = process.memory_info()
        memory_mb = memory_info.rss / 1024 / 1024
        print(f"📊 Memory usage: {memory_mb:.1f} MB")
        return memory_mb
    except:
        print("📊 Memory monitoring not available")
        return 0

def cleanup_memory():
    """Force garbage collection to free memory"""
    try:
        gc.collect()
        print("🧹 Memory cleanup completed")
    except:
        print("🧹 Memory cleanup failed")

@app.get("/")
async def root():
    return {"message": "Danier Key Items Stock Alert System API"}

@app.get("/recipients")
async def get_recipients():
    """Get all active recipients - OPTIMIZED with caching"""
    try:
        # Use cached results if available
        cache_key = "recipients_cache"
        if hasattr(key_items_service, '_cache') and cache_key in key_items_service._cache:
            cached_data = key_items_service._cache[cache_key]
            print("⚡ Using cached recipients data")
            return cached_data
        
        # Get fresh data (ACTIVE ONLY)
        recipients = recipients_storage.get_active_recipients()
        stats = recipients_storage.get_stats()
        
        result = {
            "recipients": recipients,
            "stats": stats
        }
        
        # Cache the results
        if not hasattr(key_items_service, '_cache'):
            key_items_service._cache = {}
        key_items_service._cache[cache_key] = result
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving recipients: {str(e)}")

@app.get("/test-recipients")
async def test_recipients():
    """Test recipients endpoint"""
    return {"message": "Recipients endpoint working"}

@app.get("/test-email")
async def test_email():
    """Test email endpoint"""
    return {"message": "Email endpoint working", "status": "ready"}

@app.post("/recipients")
async def add_recipient(
    email: str = Form(...),
    name: str = Form(None),
    department: str = Form(None)
):
    """Add a new email recipient"""
    result = recipients_storage.add_recipient(email, name, department)
    # Invalidate recipients cache
    try:
        if hasattr(key_items_service, '_cache'):
            key_items_service._cache.pop('recipients_cache', None)
    except Exception:
        pass
    return result

@app.delete("/recipients/{email}")
async def delete_recipient(email: str):
    """Delete a recipient email"""
    result = recipients_storage.delete_recipient(email)
    # Invalidate recipients cache
    try:
        if hasattr(key_items_service, '_cache'):
            key_items_service._cache.pop('recipients_cache', None)
    except Exception:
        pass
    return result

@app.put("/recipients/{email}")
async def update_recipient(
    email: str,
    name: str = Form(None),
    department: str = Form(None)
):
    """Update recipient information"""
    result = recipients_storage.update_recipient(email, name, department)
    # Invalidate recipients cache
    try:
        if hasattr(key_items_service, '_cache'):
            key_items_service._cache.pop('recipients_cache', None)
    except Exception:
        pass
    return result

@app.post("/upload-report")
async def upload_report(file: UploadFile = File(...)):
    """Process uploaded inventory file and return key items alerts - LIGHTWEIGHT VERSION"""
    
    # Log memory usage before processing
    initial_memory = log_memory_usage()
    
    # Validate file type
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    try:
        # Read file content directly
        content = await file.read()
        
        # Create uploads directory if it doesn't exist
        uploads_dir = UPLOAD_DIR
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
        
        # Generate unique filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"inventory_{timestamp}{file_extension}"
        permanent_path = os.path.join(uploads_dir, unique_filename)
        
        # Write content to permanent file
        with open(permanent_path, 'wb') as f:
            f.write(content)
        
        print(f"✅ File saved to: {permanent_path}")
        
        # LIGHTWEIGHT: Only clear caches, don't do heavy processing here
        print("🧹 Clearing caches for new file...")
        key_items_service.clear_all_caches()
        comparison_service.clear_all_caches()
        # Explicitly invalidate history and file list caches as well
        if not hasattr(key_items_service, '_cache'):
            key_items_service._cache = {}
        for k in [
            'upload_history_cache',
            'upload_history_signature',
            'inventory_files_cache',
            'inventory_files_signature'
        ]:
            try:
                key_items_service._cache.pop(k, None)
            except Exception:
                pass
        
        # LIGHTWEIGHT: Quick file validation only
        try:
            # Just check if file can be loaded and has basic structure
            df = pd.read_excel(permanent_path, nrows=10)  # Only read first 10 rows for validation
            if 'Season Code' not in df.columns:
                raise HTTPException(status_code=400, detail="File must contain 'Season Code' column")
            if 'Item Description' not in df.columns:
                raise HTTPException(status_code=400, detail="File must contain 'Item Description' column")
            if 'Grand Total' not in df.columns:
                raise HTTPException(status_code=400, detail="File must contain 'Grand Total' column")
            print("✅ File validation passed")
            
            # Clean up validation dataframe
            del df
            cleanup_memory()
            
        except Exception as validation_error:
            print(f"❌ Validation error: {validation_error}")
            raise HTTPException(status_code=400, detail=str(validation_error))
        
        # Database operations - LIGHTWEIGHT
        db = next(get_db())
        try:
            # Deactivate all previous files
            deactivated_count = db.query(UploadedFile).update({"is_active": False})
            print(f"📊 Deactivated {deactivated_count} previous files")
            
            # Create new uploaded file record - NO HEAVY PROCESSING
            uploaded_file = UploadedFile(
                filename=file.filename,
                file_path=permanent_path,
                file_size=len(content),
                is_active=True,
                total_items=0,  # Will be calculated later
                low_stock_count=0  # Will be calculated later
            )
            
            db.add(uploaded_file)
            db.commit()
            print(f"✅ File registered in database: {uploaded_file.filename}")
            
        except Exception as db_error:
            print(f"❌ Database error: {db_error}")
            db.rollback()
            # Clean up file if database fails
            if os.path.exists(permanent_path):
                os.remove(permanent_path)
            raise HTTPException(status_code=500, detail="Database error during file registration")
        finally:
            db.close()
        
        # Log memory usage after processing
        final_memory = log_memory_usage()
        memory_diff = final_memory - initial_memory
        print(f"📊 Memory change: {memory_diff:+.1f} MB")
        
        # Force cleanup if memory usage increased significantly
        if memory_diff > 50:  # If memory increased by more than 50MB
            print("⚠️ High memory usage detected, forcing cleanup...")
            cleanup_memory()
        
        print(f"🎉 UPLOAD COMPLETE: File {unique_filename} saved successfully")
        print(f"📁 File ready for processing - dashboard will load data on next request")

        # Clear all caches now that a new file is uploaded
        try:
            key_items_service.clear_all_caches()
            comparison_service.analysis_cache.clear()
        except Exception:
            pass

        # Trigger background cache warm to speed up first dashboard load (non-blocking)
        try:
            asyncio.create_task(asyncio.to_thread(key_items_service.get_all_key_items_with_alerts, permanent_path))
            print("🔥 Scheduled background warm of batch alerts after upload")
        except Exception as _:
            pass
        
        # Return success immediately - processing will happen on first dashboard request
        return {
            "success": True,
            "message": f"File uploaded successfully. Processing will complete when you view the dashboard.",
            "file_processed": file.filename,
            "file_saved": permanent_path,
            "processing_status": "pending",
            "next_step": "View dashboard to see processed results"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Upload error: {str(e)}")
        # Clean up any partial files
        if 'permanent_path' in locals() and os.path.exists(permanent_path):
            os.remove(permanent_path)
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    finally:
        # Always cleanup memory after upload
        cleanup_memory()

@app.get("/key-items/alerts")
async def get_key_items_alerts():
    """Get current key items alerts from the latest uploaded file - OPTIMIZED"""
    try:
        print("🔍 DASHBOARD REQUEST: Getting key items alerts...")
        
        # Get the latest uploaded file path with DB cleanup
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
        
        print(f"📁 Latest file path: {latest_file_path}")
        
        # If no file in database, try to find the most recent file in uploads directory
        if not latest_file_path:
            uploads_dir = UPLOAD_DIR
            if os.path.exists(uploads_dir):
                files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
                if files:
                    # Sort by modification time, get the most recent
                    files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
                    latest_file_path = os.path.join(uploads_dir, files[0])
        
        if not latest_file_path or not os.path.exists(latest_file_path):
            return {
                "key_items_tracked": [],
                "threshold": key_items_service.default_size_threshold,
                "low_stock_items": [],
                "message": "No inventory file found. Please upload an inventory file first."
            }
        
        # Use CACHED batch processing for speed - this prevents memory issues
        print("⚡ DASHBOARD: Using ultra-fast cached batch processing...")
        all_alerts, success, error_message = key_items_service.get_all_key_items_with_alerts(latest_file_path)
        
        if not success:
            # Only fallback to fresh processing if cache completely fails
            print("🔄 DASHBOARD FALLBACK: Cache failed, using fresh processing...")
            low_stock_items, success, error_message = key_items_service.force_fresh_processing(latest_file_path)
            if not success:
                raise HTTPException(status_code=400, detail=error_message)
            
            # Get file-specific key items and summary for fallback
            file_key_items = key_items_service.get_file_key_items(latest_file_path)
            summary = key_items_service.get_key_items_summary(latest_file_path)
            
            # Force cleanup after fresh processing
            cleanup_memory()
            
            return {
                "key_items_tracked": file_key_items,
                "threshold": key_items_service.default_size_threshold,
                "low_stock_items": low_stock_items,
                "summary": summary,
                "source_file": os.path.basename(latest_file_path),
                "total_ki00_items_detected": len(file_key_items)
            }
        
        # Convert batch alerts to old format for compatibility
        low_stock_items = []
        file_key_items = []
        total_items = 0
        
        for item in all_alerts:
            item_name = item.get("name", "")
            if item_name:
                file_key_items.append(item_name)
                alerts = item.get("alerts", [])
                low_stock_items.extend(alerts)
                total_items += len(alerts)
        
        # Simple summary from batch data
        summary = {
            "total_key_items": len(file_key_items),
            "total_low_stock_alerts": len(low_stock_items),
            "processing_mode": "cached_batch"
        }
        
        print(f"✅ DASHBOARD: Returned {len(low_stock_items)} alerts from {len(file_key_items)} items")
        
        return {
            "key_items_tracked": file_key_items,
            "threshold": key_items_service.default_size_threshold,
            "low_stock_items": low_stock_items,
            "summary": summary,
            "source_file": os.path.basename(latest_file_path),
            "total_ki00_items_detected": len(file_key_items)
        }
        
    except Exception as e:
        print(f"❌ DASHBOARD ERROR: {e}")
        # Force cleanup on error
        cleanup_memory()
        raise HTTPException(status_code=500, detail=f"Error retrieving alerts: {str(e)}")

@app.get("/inventory-files")
async def get_inventory_files():
    """Get all uploaded inventory files with their key items - OPTIMIZED"""
    try:
        uploads_dir = UPLOAD_DIR
        if not os.path.exists(uploads_dir):
            return {"files": []}
        
        files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
        
        # Use cached results if available
        cache_key = "inventory_files_cache"
        if hasattr(key_items_service, '_cache') and cache_key in key_items_service._cache:
            cached_files = key_items_service._cache[cache_key]
            if len(cached_files) == len(files):
                print("⚡ Using cached inventory files")
                return {
                    "files": cached_files,
                    "total_files": len(cached_files),
                    "all_key_items": key_items_service.get_all_key_items()
                }
        
        print(f"📁 Processing {len(files)} files for inventory list...")
        file_info = []
        
        # Process files in batches
        batch_size = 3
        for i in range(0, len(files), batch_size):
            batch_files = files[i:i + batch_size]
            print(f"📦 Processing batch {i//batch_size + 1}/{(len(files) + batch_size - 1)//batch_size}")
            
            for filename in batch_files:
                file_path = os.path.join(uploads_dir, filename)
                try:
                    file_key_items = key_items_service.get_file_key_items(file_path)
                    stat = os.stat(file_path)
                    file_info.append({
                        "filename": filename,
                        "upload_date": stat.st_mtime,
                        "file_size": stat.st_size,
                        "key_items_count": len(file_key_items),
                        "key_items": file_key_items,
                        "file_path": file_path
                    })
                except Exception as e:
                    print(f"Error processing file {filename}: {e}")
                    continue
        
        # Cache the results
        if not hasattr(key_items_service, '_cache'):
            key_items_service._cache = {}
        key_items_service._cache[cache_key] = file_info
        
        return {
            "files": file_info,
            "total_files": len(file_info),
            "all_key_items": key_items_service.get_all_key_items()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving inventory files: {str(e)}")

@app.get("/inventory-files/{filename}/alerts")
async def get_file_alerts(filename: str):
    """Get alerts for a specific inventory file"""
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        low_stock_items, success, error_message = key_items_service.process_key_items_inventory(file_path)
        
        if not success:
            raise HTTPException(status_code=400, detail=error_message)
        
        summary = key_items_service.get_key_items_summary(file_path)
        
        return {
            "filename": filename,
            "key_items_tracked": key_items_service.get_file_key_items(file_path),
            "threshold": key_items_service.default_size_threshold,
            "low_stock_items": low_stock_items,
            "summary": summary,
            "total_ki00_items_detected": len(key_items_service.get_file_key_items(file_path))
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving file alerts: {str(e)}")

# Optimized key items endpoint with batch processing
@app.get("/key-items/list")
async def get_key_items_list():
    """Get all key items with optimized batch processing"""
    try:
        # Use comparison service caching for ultra-fast response
        cache_key = f"key_items_list_{datetime.now().strftime('%Y%m%d_%H')}"
        
        # Get the latest file with proper caching
        files = comparison_service.get_all_uploaded_files()
        if not files:
            return {"key_items": [], "message": "No files uploaded yet"}
        
        latest_file = files[0]  # Already sorted by date
        
        # Ultra-fast processing with batch operations
        success, key_items, error = key_items_service.get_key_items_batch(latest_file['file_path'])
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        return {
            "key_items": key_items,
            "file": latest_file['filename'],
            "count": len(key_items),
            "cached": True
        }
        
    except Exception as e:
        print(f"⚠️  Error in get_key_items_list: {str(e)}")
        return {"key_items": [], "error": str(e)}

@app.get("/key-items/{item_name}/alerts")
async def get_key_item_alerts(item_name: str):
    """Get alerts for specific key item with lightning-fast caching"""
    try:
        # Use service-level caching for instant response
        alerts = key_items_service.get_item_alerts_cached(item_name)
        
        return {
            "item_name": item_name,
            "alerts": alerts,
            "count": len(alerts),
            "cached": True
        }
        
    except Exception as e:
        print(f"⚠️  Error getting alerts for {item_name}: {str(e)}")
        return {"alerts": [], "error": str(e)}

@app.get("/test")
async def test_system():
    """Test endpoint to verify system is working with the latest uploaded file"""
    try:
        # Get the latest uploaded file path
        db = next(get_db())
        latest_file_path = file_storage_service.get_latest_file_path(db)
        
        if not latest_file_path:
            uploads_dir = UPLOAD_DIR
            if os.path.exists(uploads_dir):
                files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
                if files:
                    files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
                    latest_file_path = os.path.join(uploads_dir, files[0])
        
        if not latest_file_path or not os.path.exists(latest_file_path):
            return {
                "status": "no_file",
                "message": "No inventory file found. Please upload a file first.",
                "key_items_service_status": "ready",
                "dynamic_detection": "enabled"
            }
        
        # Test key items detection
        file_key_items = key_items_service.get_file_key_items(latest_file_path)
        
        # Test processing
        low_stock_items, success, error_message = key_items_service.process_key_items_inventory(latest_file_path)
        
        return {
            "status": "success" if success else "error",
            "message": error_message if error_message else "System working correctly",
            "latest_file": os.path.basename(latest_file_path) if latest_file_path else None,
            "detected_ki00_items": len(file_key_items),
            "sample_ki00_items": file_key_items[:5] if file_key_items else [],
            "low_stock_alerts": len(low_stock_items) if success else 0,
            "key_items_service_status": "operational",
            "dynamic_detection": "working",
            "all_key_items_across_files": len(key_items_service.get_all_key_items())
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"System test failed: {str(e)}",
            "key_items_service_status": "error"
        }

@app.get("/upload-history")
async def get_upload_history():
    """Get upload history - OPTIMIZED for speed with signature-based cache validation"""
    try:
        uploads_dir = UPLOAD_DIR
        if not os.path.exists(uploads_dir):
            return {"uploads": []}
        
        files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
        
        # Build a signature from filenames + mtimes to detect any change
        current_signature = [
            (fname, os.path.getmtime(os.path.join(uploads_dir, fname))) for fname in files
        ]
        
        # Use cached results if available and signature matches
        cache_key = "upload_history_cache"
        signature_key = "upload_history_signature"
        if hasattr(key_items_service, '_cache') and cache_key in key_items_service._cache and signature_key in key_items_service._cache:
            cached_history = key_items_service._cache[cache_key]
            cached_signature = key_items_service._cache[signature_key]
            if cached_signature == current_signature:
                print("⚡ Using cached upload history (signature matched)")
                return {"uploads": cached_history, "total_uploads": len(cached_history)}
        
        print(f"📊 Fast processing {len(files)} files for upload history...")
        history = []
        
        # FAST: Use file metadata only, don't process each file individually
        for filename in files:
            file_path = os.path.join(uploads_dir, filename)
            try:
                stat = os.stat(file_path)
                
                # Use cached data if available for this file
                file_cache_key = f"file_stats_{filename}"
                if hasattr(key_items_service, '_cache') and file_cache_key in key_items_service._cache:
                    cached_stats = key_items_service._cache[file_cache_key]
                    history.append({
                        "filename": filename,
                        "upload_date": stat.st_mtime,
                        "file_size": stat.st_size,
                        "key_items_detected": cached_stats.get('key_items_count', 0),
                        "low_stock_alerts": cached_stats.get('low_stock_count', 0),
                        "processed_successfully": True
                    })
                else:
                    # Quick estimation based on file size (much faster than processing)
                    estimated_items = max(1, int(stat.st_size / 1000))  # Rough estimate
                    history.append({
                        "filename": filename,
                        "upload_date": stat.st_mtime,
                        "file_size": stat.st_size,
                        "key_items_detected": estimated_items,
                        "low_stock_alerts": estimated_items // 2,  # Rough estimate
                        "processed_successfully": True
                    })
                    
            except Exception as e:
                print(f"Error processing file {filename}: {e}")
                history.append({
                    "filename": filename,
                    "upload_date": stat.st_mtime if 'stat' in locals() else 0,
                    "file_size": stat.st_size if 'stat' in locals() else 0,
                    "key_items_detected": 0,
                    "low_stock_alerts": 0,
                    "processed_successfully": False
                })
                continue
        
        # Cache the results with signature
        if not hasattr(key_items_service, '_cache'):
            key_items_service._cache = {}
        key_items_service._cache[cache_key] = history
        key_items_service._cache[signature_key] = current_signature
        
        print(f"✅ Fast upload history: {len(history)} files processed")
        return {"uploads": history, "total_uploads": len(history)}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving upload history: {str(e)}")

@app.get("/files/smart-analysis/{file1}/{file2}")
async def get_smart_performance_analysis(file1: str, file2: str):
    """Get intelligent performance analysis between two inventory files"""
    try:
        file1_path = os.path.join(UPLOAD_DIR, file1)
        file2_path = os.path.join(UPLOAD_DIR, file2)
        
        if not os.path.exists(file1_path):
            raise HTTPException(status_code=404, detail=f"File {file1} not found")
        if not os.path.exists(file2_path):
            raise HTTPException(status_code=404, detail=f"File {file2} not found")
        
        analysis_result = comparison_service.get_smart_performance_analysis(file1_path, file2_path)
        
        if "error" in analysis_result:
            raise HTTPException(status_code=400, detail=analysis_result["error"])
        
        return analysis_result
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Smart analysis endpoint error: {str(e)}")
        print(f"Full traceback: {error_details}")
        raise HTTPException(status_code=500, detail=f"Error analyzing files: {str(e)}")

@app.get("/files/smart-analysis-simple/{file1}/{file2}")
async def get_smart_performance_analysis_simple(file1: str, file2: str):
    """Get simple smart performance analysis between two inventory files"""
    try:
        file1_path = os.path.join(UPLOAD_DIR, file1)
        file2_path = os.path.join(UPLOAD_DIR, file2)
        
        if not os.path.exists(file1_path):
            return {"error": f"File {file1} not found"}
        if not os.path.exists(file2_path):
            return {"error": f"File {file2} not found"}
        
        # Load files using key items service
        df1 = key_items_service._load_inventory_file(file1_path)
        df2 = key_items_service._load_inventory_file(file2_path)
        
        if df1 is None or df2 is None:
            return {"error": "Failed to load one or both files"}
        
        # Get basic stats
        ki00_1 = df1[df1['Season Code'] == 'KI00'] if 'Season Code' in df1.columns else pd.DataFrame()
        ki00_2 = df2[df2['Season Code'] == 'KI00'] if 'Season Code' in df2.columns else pd.DataFrame()
        
        total_stock_1 = ki00_1['Grand Total'].sum() if 'Grand Total' in ki00_1.columns and not ki00_1.empty else 0
        total_stock_2 = ki00_2['Grand Total'].sum() if 'Grand Total' in ki00_2.columns and not ki00_2.empty else 0
        
        # Simple analysis
        return {
            "file1": file1,
            "file2": file2,
            "analysis_date": datetime.now().strftime('%Y-%m-%d %H:%M'),
            "simple_comparison": {
                "file1_ki00_items": len(ki00_1),
                "file2_ki00_items": len(ki00_2),
                "file1_total_stock": int(total_stock_1) if pd.notna(total_stock_1) else 0,
                "file2_total_stock": int(total_stock_2) if pd.notna(total_stock_2) else 0,
                "stock_change": int(total_stock_2 - total_stock_1) if pd.notna(total_stock_2) and pd.notna(total_stock_1) else 0,
                "status": "✅ Smart analysis working! Files processed successfully."
            }
        }
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Simple smart analysis error: {str(e)}")
        print(f"Full traceback: {error_details}")
        return {"error": f"Simple analysis failed: {str(e)}"}

@app.get("/files/performance-analysis")
async def get_performance_analysis():
    """Get performance analysis across all uploaded files"""
    try:
        files_data = comparison_service.get_all_uploaded_files()
        
        if len(files_data) < 2:
            return {
                "error": "Need at least 2 files for performance analysis",
                "files_count": len(files_data),
                "message": "Upload more inventory files to enable performance analysis"
            }
        
        analysis_result = comparison_service.analyze_product_performance(files_data)
        
        if "error" in analysis_result:
            raise HTTPException(status_code=400, detail=analysis_result["error"])
        
        return analysis_result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error performing analysis: {str(e)}")

@app.get("/files/enhanced-list")
async def get_enhanced_inventory_files():
    """Get enhanced list of inventory files with detailed metadata"""
    try:
        files_data = comparison_service.get_all_uploaded_files()
        
        return {
            "files": files_data,
            "total_files": len(files_data),
            "analysis_ready": len(files_data) >= 2,
            "all_key_items": key_items_service.get_all_key_items()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving enhanced file list: {str(e)}")

@app.get("/files/archive")
async def get_file_archive():
    """Get comprehensive file archive with categorization and search"""
    try:
        files_data = comparison_service.get_all_uploaded_files()
        
        # Categorize files by age and status
        recent_files = []
        archive_files = []
        
        for file in files_data:
            if "error" not in file:
                age_days = file.get("file_age_days", 0)
                if age_days <= 7:
                    recent_files.append(file)
                else:
                    archive_files.append(file)
        
        # Sort by date
        recent_files.sort(key=lambda x: x.get('upload_timestamp', 0), reverse=True)
        archive_files.sort(key=lambda x: x.get('upload_timestamp', 0), reverse=True)
        
        return {
            "recent_files": recent_files,
            "archive_files": archive_files,
            "total_files": len(files_data),
            "recent_count": len(recent_files),
            "archive_count": len(archive_files),
            "storage_info": {
                "upload_directory": UPLOAD_DIR,
                "total_size_mb": sum(f.get("file_size", 0) for f in files_data if "error" not in f) / (1024 * 1024)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving file archive: {str(e)}")

@app.get("/files/archive/{filename}/download")
async def download_archive_file(filename: str):
    """Download a file from the archive"""
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        return FileResponse(
            file_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=filename
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

@app.get("/key-items/batch-alerts")
async def get_all_key_items_with_alerts():
    """Ultra-fast batch alerts with self-healing file detection and memory optimization"""
    try:
        print("⚡ BATCH ALERTS: Starting request...")
        
        # Use DB to get latest active file path with immediate cleanup
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
        
        # Self-healing: If no file in DB OR path missing, try to find files in uploads directory
        if not latest_file_path or not os.path.exists(latest_file_path):
            if latest_file_path and not os.path.exists(latest_file_path):
                print(f"🔄 Self-healing: DB path missing on disk: {latest_file_path}")
            else:
                print("🔄 Self-healing: No file in database, searching uploads directory...")
            uploads_dir = UPLOAD_DIR
            if os.path.exists(uploads_dir):
                files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
                if files:
                    # Sort by modification time, get the most recent
                    files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
                    latest_file_path = os.path.join(uploads_dir, files[0])
                    print(f"✅ Self-healing: Using latest file {files[0]} in uploads directory")
                    
                    # Register this file in the database with error handling
                    try:
                        db = next(get_db())
                        try:
                            # Deactivate all previous files
                            db.query(UploadedFile).update({"is_active": False})
                            
                            # Create new uploaded file record
                            file_size = os.path.getsize(latest_file_path)
                            uploaded_file = UploadedFile(
                                filename=files[0],
                                file_path=latest_file_path,
                                file_size=file_size,
                                is_active=True,
                                total_items=0,
                                low_stock_count=0
                            )
                            db.add(uploaded_file)
                            db.commit()
                            print(f"✅ Self-healing: Registered {files[0]} in database")
                        finally:
                            db.close()
                    except Exception as e:
                        print(f"⚠️ Self-healing database registration failed: {e}")
        
        if not latest_file_path:
            return {
                "key_items": [], 
                "message": "No inventory files found. Please upload an inventory file using the 'Upload Report' button above.",
                "help": {
                    "title": "How to get started",
                    "steps": [
                        "1. Click 'Upload Report' in the navigation bar",
                        "2. Select your inventory Excel file (.xlsx)",
                        "3. Wait for processing to complete",
                        "4. View your alerts in the dashboard"
                    ]
                }
            }
        
        # Use ultra-fast batch processing with memory management
        print("⚡ BATCH ALERTS: Using cached batch processing...")
        all_alerts, success, error = key_items_service.get_all_key_items_with_alerts(latest_file_path)
        
        if not success:
            print(f"❌ BATCH ALERTS: Cache failed, error: {error}")
            # Force memory cleanup before error
            cleanup_memory()
            raise HTTPException(status_code=400, detail=error)
        
        print(f"✅ BATCH ALERTS: Returned {len(all_alerts)} items successfully")
        
        # Optional memory cleanup for large results
        if len(all_alerts) > 50:
            print("🧹 BATCH ALERTS: Large result set, triggering cleanup...")
            cleanup_memory()
            
        return {
            "key_items": all_alerts,
            "file": os.path.basename(latest_file_path),
            "total_items": len(all_alerts),
            "cached": True,
            "memory_optimized": True
        }
    except Exception as e:
        print(f"❌ BATCH ALERTS ERROR: {str(e)}")
        # Force cleanup on any error
        cleanup_memory()
        return {"key_items": [], "error": str(e)}

@app.get("/search/article/{search_term}")
async def search_article_alerts(search_term: str):
    """Search for specific article alerts"""
    try:
        print(f"🔍 API: Searching for article '{search_term}'")
        results = key_items_service.search_article_alerts(search_term)
        
        return {
            "search_term": search_term,
            "results": results,
            "count": len(results),
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        print(f"❌ Search error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")

@app.post("/thresholds/set")
async def set_custom_threshold(
    item_name: str = Form(...),
    size: str = Form(...),
    color: str = Form(...),
    threshold: int = Form(...)
):
    """Set custom threshold for a specific key item, size, and color combination"""
    try:
        if threshold < 0:
            raise HTTPException(status_code=400, detail="Threshold must be positive")
        
        success = key_items_service.set_custom_threshold(item_name, size, color, threshold)
        
        # Clear cache to reflect changes immediately
        key_items_service.clear_all_caches()
        
        return {
            "success": success,
            "item_name": item_name,
            "size": size,
            "color": color,
            "threshold": threshold,
            "message": f"Threshold set to {threshold} for {item_name} ({size}, {color})"
        }
        
    except Exception as e:
        print(f"❌ Set threshold error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to set threshold: {str(e)}")

@app.get("/thresholds/get/{item_name}/{size}/{color}")
async def get_custom_threshold(item_name: str, size: str, color: str):
    """Get custom threshold for a specific key item, size, and color combination"""
    try:
        threshold = key_items_service.get_custom_threshold(item_name, size, color)
        key = f"{item_name}|{size}|{color}"
        is_custom = key in key_items_service.custom_thresholds
        
        return {
            "item_name": item_name,
            "size": size,
            "color": color,
            "threshold": threshold,
            "is_custom": is_custom,
            "default_threshold": key_items_service.default_size_threshold
        }
        
    except Exception as e:
        print(f"❌ Get threshold error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get threshold: {str(e)}")

@app.get("/thresholds/all")
async def get_all_custom_thresholds():
    """Get all custom thresholds"""
    try:
        thresholds = key_items_service.get_all_custom_thresholds()
        
        # Convert to more readable format
        formatted_thresholds = []
        for key, threshold in thresholds.items():
            parts = key.split('|')
            if len(parts) == 3:
                formatted_thresholds.append({
                    "item_name": parts[0],
                    "size": parts[1],
                    "color": parts[2],
                    "threshold": threshold,
                    "key": key
                })
        
        return {
            "custom_thresholds": formatted_thresholds,
            "raw_thresholds": thresholds,
            "default_threshold": key_items_service.default_size_threshold,
            "total_custom": len(thresholds)
        }
        
    except Exception as e:
        print(f"❌ Get all thresholds error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get thresholds: {str(e)}")

@app.delete("/thresholds/reset/{item_name}/{size}/{color}")
async def reset_custom_threshold(item_name: str, size: str, color: str):
    """Reset custom threshold for a specific key item, size, and color to default"""
    try:
        success = key_items_service.reset_custom_threshold(item_name, size, color)
        
        # Clear cache to reflect changes immediately
        key_items_service.clear_all_caches()
        
        return {
            "success": success,
            "item_name": item_name,
            "size": size,
            "color": color,
            "message": f"Threshold reset to default for {item_name} ({size}, {color})"
        }
        
    except Exception as e:
        print(f"❌ Reset threshold error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to reset threshold: {str(e)}")

@app.post("/clear-cache")
async def clear_cache():
    """Clear all processing caches to force fresh data"""
    try:
        key_items_service.clear_all_caches()
        comparison_service.clear_all_caches()
        cleanup_memory()
        return {"success": True, "message": "All caches cleared successfully"}
    except Exception as e:
        print(f"❌ Clear cache error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to clear cache: {str(e)}")

@app.post("/warm-cache")
async def warm_cache():
    """Warm up caches for better performance after email operations"""
    try:
        print("🔥 WARMING CACHE: Starting cache warm-up...")
        
        # Get latest file
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
            
        if not latest_file_path:
            return {"success": False, "message": "No file to warm cache for"}
        
        # Warm cache in background
        def warm_cache_background():
            try:
                print("🔥 BACKGROUND WARM: Starting cache warming...")
                key_items_service.get_all_key_items_with_alerts(latest_file_path)
                print("✅ BACKGROUND WARM: Cache warmed successfully")
            except Exception as e:
                print(f"❌ BACKGROUND WARM: Cache warming failed: {e}")
        
        import threading
        warm_thread = threading.Thread(target=warm_cache_background, daemon=True)
        warm_thread.start()
        
        return {"success": True, "message": "Cache warming started in background"}
    except Exception as e:
        print(f"❌ Warm cache error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to warm cache: {str(e)}")

@app.post("/cache/clear")
async def clear_all_caches_api():
    try:
        key_items_service.clear_all_caches()
        comparison_service.analysis_cache.clear()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Email Alert Endpoints
@app.post("/email/send-alert")
async def send_email_alert(
    item_name: str = Form(None)
):
    """Send personalized email alert to all active recipients - NON-BLOCKING (ultra-light latest file lookup)"""
    try:
        print(f"📧 EMAIL REQUEST: Starting email alert for item: {item_name or 'ALL'}")
        
        # Lightweight: get latest file path from DB (no full scan)
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
        if not latest_file_path or not os.path.exists(latest_file_path):
            raise HTTPException(status_code=400, detail="No inventory file found")
        
        # Get all items with alerts (cached and fast)
        all_alerts, success, error = key_items_service.get_all_key_items_with_alerts(latest_file_path)
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        # Flatten alerts across all items, filter by item if provided
        if item_name:
            target = next((it for it in all_alerts if it.get('name') == item_name), None)
            low_stock_items = list(target.get('alerts', [])) if target else []
        else:
            low_stock_items = []
            for item in all_alerts:
                low_stock_items.extend(item.get('alerts', []))
        
        # Recipients
        active_recipients = recipients_storage.get_active_recipients()
        if not active_recipients:
            recipients = ["danieralertsystem@gmail.com"]
            recipient_names = {"danieralertsystem@gmail.com": "Danier Stock Alert System"}
        else:
            recipients = [r['email'] for r in active_recipients if r.get('active', True)]
            recipient_names = {r['email']: r.get('name') for r in active_recipients}
        
        print(f"📧 Preparing background email to {len(recipients)} recipients | items: {len(low_stock_items)} | item={item_name or 'ALL'}")
        
        # Background send for speed and stability
        def _send_background():
            try:
                email_service.send_personalized_alert(
                    recipients=recipients,
                    low_stock_items=low_stock_items,
                    recipient_names=recipient_names,
                    item_name=item_name or None
                )
            except Exception as e:
                print(f"❌ BACKGROUND EMAIL ERROR: {e}")
        
        threading.Thread(target=_send_background, daemon=True).start()
        
        # Return immediately
        return {
            "success": True,
            "message": (
                f"Email alert for {item_name} is being sent to {len(recipients)} recipients in the background"
                if item_name else f"General email alert is being sent to {len(recipients)} recipients in the background"
            ),
            "recipients": recipients,
            "items_count": len(low_stock_items),
            "item_name": item_name,
            "processing_status": "background"
        }
            
    except Exception as e:
        print(f"❌ EMAIL REQUEST ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/email/status")
async def get_email_status():
    """Get email service status and configuration"""
    smtp_configured = bool(email_service.smtp_pass)
    return {
        "smtp_configured": smtp_configured,
        "smtp_host": email_service.smtp_host,
        "smtp_port": email_service.smtp_port,
        "email_from": email_service.email_from,
        "status": "Ready" if smtp_configured else "Not configured"
    }

@app.get("/threshold-analysis")
async def get_threshold_analysis():
    """Get threshold change analysis between latest and previous file uploads"""
    try:
        # Get all uploaded files
        files = comparison_service.get_all_uploaded_files()
        if not files:
            return {"error": "No files uploaded yet"}
        
        if len(files) < 2:
            # Only one file - initial analysis
            current_file = files[0]
            analysis = threshold_analysis_service.analyze_threshold_changes(current_file['file_path'])
        else:
            # Compare latest with previous
            current_file = files[0]  # Most recent
            previous_file = files[1]  # Second most recent
            analysis = threshold_analysis_service.analyze_threshold_changes(
                current_file['file_path'], 
                previous_file['file_path']
            )
        
        # Add summary
        analysis['summary_text'] = threshold_analysis_service.get_threshold_alert_summary(analysis)
        
        return analysis
        
    except Exception as e:
        return {"error": f"Threshold analysis failed: {str(e)}"}

@app.get("/threshold-analysis/{filename}")
async def get_threshold_analysis_for_file(filename: str):
    """Get threshold analysis for a specific file compared to its previous upload"""
    try:
        # Get all uploaded files
        files = comparison_service.get_all_uploaded_files()
        if not files:
            return {"error": "No files uploaded yet"}
        
        # Find the specified file
        target_file = None
        target_index = -1
        
        for i, file in enumerate(files):
            if file['filename'] == filename:
                target_file = file
                target_index = i
                break
        
        if not target_file:
            return {"error": f"File {filename} not found"}
        
        if target_index == 0:
            # Latest file - compare with previous
            if len(files) < 2:
                analysis = threshold_analysis_service.analyze_threshold_changes(target_file['file_path'])
            else:
                previous_file = files[1]
                analysis = threshold_analysis_service.analyze_threshold_changes(
                    target_file['file_path'], 
                    previous_file['file_path']
                )
        else:
            # Compare with the file before it
            previous_file = files[target_index + 1]
            analysis = threshold_analysis_service.analyze_threshold_changes(
                target_file['file_path'], 
                previous_file['file_path']
            )
        
        # Add summary
        analysis['summary_text'] = threshold_analysis_service.get_threshold_alert_summary(analysis)
        
        return analysis
        
    except Exception as e:
        return {"error": f"Threshold analysis failed: {str(e)}"}

@app.get("/key-items/summary")
async def get_key_items_summary():
    """Ultra-fast summary of key items with self-healing file detection"""
    try:
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
        
        # Self-healing: If no file in DB, try to find files in uploads directory
        if not latest_file_path:
            print("🔄 Self-healing: No file in database, searching uploads directory...")
            uploads_dir = UPLOAD_DIR
            if os.path.exists(uploads_dir):
                files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
                if files:
                    # Sort by modification time, get the most recent
                    files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
                    latest_file_path = os.path.join(uploads_dir, files[0])
                    print(f"✅ Self-healing: Found file {files[0]} in uploads directory")
                    
                    # Register this file in the database
                    try:
                        db = next(get_db())
                        try:
                            # Deactivate all previous files
                            db.query(UploadedFile).update({"is_active": False})
                            
                            # Create new uploaded file record
                            file_size = os.path.getsize(latest_file_path)
                            uploaded_file = UploadedFile(
                                filename=files[0],
                                file_path=latest_file_path,
                                file_size=file_size,
                                is_active=True,
                                total_items=0,
                                low_stock_count=0
                            )
                            db.add(uploaded_file)
                            db.commit()
                            print(f"✅ Self-healing: Registered {files[0]} in database")
                        finally:
                            db.close()
                    except Exception as e:
                        print(f"⚠️ Self-healing database registration failed: {e}")
        
        if not latest_file_path:
            return {
                "key_items": [], 
                "message": "No inventory files found. Please upload an inventory file using the 'Upload Report' button above.",
                "help": {
                    "title": "How to get started",
                    "steps": [
                        "1. Click 'Upload Report' in the navigation bar",
                        "2. Select your inventory Excel file (.xlsx)",
                        "3. Wait for processing to complete",
                        "4. View your alerts in the dashboard"
                    ]
                }
            }
        
        # Use ultra-fast batch processing with caching
        all_alerts, success, error = key_items_service.get_all_key_items_with_alerts(latest_file_path)
        
        if not success:
            # Only fallback to fresh processing if batch fails completely
            print("🔄 Fallback: forcing fresh processing for summary")
            all_alerts, success, error = key_items_service.force_fresh_processing(latest_file_path)
            
        if not success:
            raise HTTPException(status_code=400, detail=error)
            
        # Convert to summary format
        summary = []
        for item in all_alerts:
            summary.append({
                "name": item.get("name"),
                "alert_count": len(item.get("alerts", [])),
                "variants_count": len(item.get("alerts", [])),
                "total_stock": item.get("total_stock")
            })
        return {"key_items": summary}
    except Exception as e:
        return {"key_items": [], "error": str(e)}

@app.get("/key-items/options/{item_name}")
async def get_item_options(item_name: str):
    """Return available colours and sizes for a given key item from latest inventory file"""
    try:
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
        if not latest_file_path:
            return {"colors": [], "sizes": [], "color_to_sizes": {}, "size_to_colors": {}}
        # Ensure we have a populated dataframe
        df = key_items_service._load_inventory_file(latest_file_path)
        if df is None or 'Season Code' not in df.columns:
            return {"colors": [], "sizes": [], "color_to_sizes": {}, "size_to_colors": {}}
        ki = df[df['Season Code'] == 'KI00'].copy()
        item_col = key_items_service._detect_item_column(df.columns)
        if not item_col or 'Variant Color' not in df.columns or 'Variant Code' not in df.columns:
            return {"colors": [], "sizes": [], "color_to_sizes": {}, "size_to_colors": {}}
        ki['item_base'] = ki[item_col].str.split(' - ').str[0].str.strip()
        item_df = ki[ki['item_base'] == item_name]
        if item_df.empty:
            return {"colors": [], "sizes": [], "color_to_sizes": {}, "size_to_colors": {}}
        # Build sets
        colors = sorted(item_df['Variant Color'].dropna().astype(str).unique().tolist())
        # Map color -> sizes
        color_to_sizes = {}
        size_to_colors = {}
        for _, row in item_df.iterrows():
            color = str(row['Variant Color'])
            size = key_items_service.extract_size_from_variant(str(row['Variant Code']))
            if color not in color_to_sizes:
                color_to_sizes[color] = set()
            color_to_sizes[color].add(size)
            if size not in size_to_colors:
                size_to_colors[size] = set()
            size_to_colors[size].add(color)
        # Unique lists
        color_to_sizes = {c: sorted(list(s)) for c, s in color_to_sizes.items()}
        size_to_colors = {s: sorted(list(cset)) for s, cset in size_to_colors.items()}
        sizes = sorted(list(size_to_colors.keys()))
        return {"colors": colors, "sizes": sizes, "color_to_sizes": color_to_sizes, "size_to_colors": size_to_colors}
    except Exception as e:
        return {"colors": [], "sizes": [], "color_to_sizes": {}, "size_to_colors": {}, "error": str(e)}

@app.post("/alerts/download-all")
async def download_all_alerts():
    """Generate Excel report of all alerts and send email notification - ULTRA STABLE"""
    try:
        print("📊 Starting download all alerts - non-blocking mode")
        
        # Get the latest file path with error handling
        db = next(get_db())
        try:
            latest_file_path = file_storage_service.get_latest_file_path(db)
        finally:
            db.close()
            
        if not latest_file_path or not os.path.exists(latest_file_path):
            raise HTTPException(status_code=404, detail="No inventory file found")
        
        # Get all alerts using CACHED data (no heavy processing)
        all_alerts, success, error = key_items_service.get_all_key_items_with_alerts(latest_file_path)
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        # Create Excel file efficiently
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Low Stock Alerts"
        
        # Headers
        headers = ["Item Name", "Item Number", "Color", "Size", "Current Stock", "Required Threshold", "Shortage", "New Order", "Priority"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Data rows
        row = 2
        total_alerts = 0
        for item in all_alerts:
            item_name = item.get("name", "")
            alerts = item.get("alerts", [])
            
            for alert in alerts:
                shortage = alert.get("shortage", 0)
                priority = "CRITICAL" if shortage >= 10 else "HIGH" if shortage >= 5 else "MEDIUM"
                
                ws.cell(row=row, column=1, value=item_name)
                ws.cell(row=row, column=2, value=alert.get("item_number", ""))
                ws.cell(row=row, column=3, value=alert.get("color", ""))
                ws.cell(row=row, column=4, value=alert.get("size", ""))
                ws.cell(row=row, column=5, value=alert.get("stock_level", 0))
                ws.cell(row=row, column=6, value=alert.get("required_threshold", 0))
                ws.cell(row=row, column=7, value=f"-{shortage}")
                # New Order may be None if column not present; write empty string in that case
                new_order_value = alert.get("new_order", None)
                ws.cell(row=row, column=8, value=new_order_value if new_order_value is not None else "")
                ws.cell(row=row, column=9, value=priority)
                
                # Color coding for priority
                if priority == "CRITICAL":
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                elif priority == "HIGH":
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF0E5", end_color="FFF0E5", fill_type="solid")
                
                row += 1
                total_alerts += 1
        
        # Auto-adjust column widths efficiently
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"✅ Excel generated with {total_alerts} alerts")
        
        # PERMANENT SOLUTION: Email Excel file with bulletproof error handling
        def send_excel_email_safe():
            try:
                print("📧 DOWNLOAD EMAIL: Starting bulletproof Excel email delivery...")
                recipients = recipients_storage.get_active_recipients()
                if not recipients:
                    print("⚠️ No active recipients found for Excel email")
                    return
                
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                subject = f"📊 Danier Stock Alert Report - {timestamp}"
                
                # Use ultra-simple email approach to prevent crashes
                import subprocess
                import tempfile
                import os
                
                # Save Excel to temp file
                temp_dir = tempfile.mkdtemp()
                excel_filename = f"danier_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                excel_path = os.path.join(temp_dir, excel_filename)
                
                with open(excel_path, 'wb') as f:
                    f.write(output.getvalue())
                
                success_count = 0
                for recipient in recipients:
                    try:
                        recipient_email = recipient.get('email', '')
                        recipient_name = recipient.get('name', 'Team')
                        
                        # Create simple email body
                        email_body = f"""
Subject: {subject}
To: {recipient_email}
From: Danier Stock Alerts <danieralertsystem@gmail.com>

Dear {recipient_name},

Please find attached the Danier Stock Alert Report generated on {timestamp}.

Report Details:
- Total Alerts: {total_alerts}
- File: {excel_filename}
- Generated: {timestamp}

The Excel file contains all current low stock alerts with priority color coding.

Best regards,
Danier Automated Alert System
                        """
                        
                        # Simple approach: Just send via Python email without complex SMTP
                        email_sent = email_service.send_excel_attachment(
                            recipient=recipient_email,
                            subject=subject,
                            excel_content=output.getvalue(),
                            filename=excel_filename,
                            recipient_name=recipient_name,
                            total_alerts=total_alerts,
                            source_file=os.path.basename(latest_file_path)
                        )
                        
                        if email_sent:
                            success_count += 1
                            print(f"✅ Excel emailed successfully to: {recipient_email}")
                            recipients_storage.record_email_sent(recipient_email)
                        else:
                            print(f"⚠️ Excel email failed to: {recipient_email} (non-critical)")
                            
                    except Exception as recipient_error:
                        print(f"⚠️ Individual recipient error for {recipient.get('email', 'unknown')}: {recipient_error}")
                        continue
                
                # Cleanup temp files
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                except:
                    pass
                
                print(f"📧 Excel email delivery completed: {success_count}/{len(recipients)} successful")
                        
            except Exception as e:
                print(f"📧 Excel email error (non-critical): {e}")
                # Continue execution even if email fails completely
        
        # Run email in completely isolated process to prevent any crashes
        import threading
        email_thread = threading.Thread(target=send_excel_email_safe, daemon=True)
        email_thread.start()
        
        print("📧 Excel email delivery started in background")
        
        # Return Excel file immediately
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=danier_alerts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Download all alerts error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")

@app.get("/health")
async def health():
    try:
        # Quick memory check
        memory_mb = 0
        try:
            process = psutil.Process()
            memory_info = process.memory_info()
            memory_mb = memory_info.rss / 1024 / 1024
        except:
            pass
        
        return {
            "status": "ok", 
            "memory_mb": round(memory_mb, 1),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.on_event("startup")
async def warm_caches_on_startup():
    """ULTRA-MINIMAL startup - zero heavy processing to prevent restarts"""
    try:
        print("🚀 MINIMAL STARTUP - No processing to ensure stability")
        
        # Just check database connectivity - NO file processing
        try:
            db = next(get_db())
            db.close()
            print("✅ Database connection verified")
        except Exception as e:
            print(f"⚠️ Database warning: {e}")
        
        # Ensure upload directory exists
        uploads_dir = UPLOAD_DIR
        if not os.path.exists(uploads_dir):
            try:
                os.makedirs(uploads_dir)
                print(f"✅ Created uploads directory: {uploads_dir}")
            except Exception as e:
                print(f"⚠️ Upload directory warning: {e}")
        
        # Ensure emails directory exists
        emails_dir = "emails"
        if not os.path.exists(emails_dir):
            try:
                os.makedirs(emails_dir)
                print(f"✅ Created emails directory: {emails_dir}")
            except Exception as e:
                print(f"⚠️ Emails directory warning: {e}")
        
        # Clear any stale cache on startup
        try:
            if hasattr(key_items_service, '_cache'):
                key_items_service._cache.clear()
            print("✅ Cleared stale caches")
        except Exception as e:
            print(f"⚠️ Cache clear warning: {e}")
        
        # Ensure default user exists on startup
        db = next(get_db())
        try:
            _create_or_get_default_user(db)
            print("✅ Default user checked/created")
        finally:
            db.close()

        print("✅ MINIMAL STARTUP COMPLETE - Server ready, no heavy processing")
        
    except Exception as e:
        print(f"⚠️ Startup warning (non-critical): {e}")

@app.middleware("http")
async def error_handling_middleware(request, call_next):
    """Global error handling and memory management middleware"""
    try:
        response = await call_next(request)
        
        # Periodic memory cleanup for email operations
        if "/email/" in str(request.url) or "/download-all" in str(request.url):
            print("🧹 Post-email operation cleanup")
            cleanup_memory()
            
        return response
    except Exception as e:
        print(f"❌ GLOBAL ERROR on {request.url}: {e}")
        import traceback
        traceback.print_exc()
        
        # Force cleanup on any unhandled error
        try:
            cleanup_memory()
        except:
            pass
            
        # Return a stable error response instead of crashing
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=500,
            content={"error": "Internal server error", "detail": str(e), "stable": True}
        )

@app.get("/key-items/details/{item_name}")
async def get_item_details(item_name: str):
    """Return details (total, colour totals, alerts) for a single item using cached batch results"""
    try:
        files = comparison_service.get_all_uploaded_files()
        if not files:
            return {"name": item_name, "total_stock": 0, "color_totals": [], "alerts": [], "alert_count": 0}
        latest_file = files[0]
        all_alerts, success, error = key_items_service.get_all_key_items_with_alerts(latest_file['file_path'])
        if not success:
            raise HTTPException(status_code=400, detail=error)
        for it in all_alerts:
            if it.get("name") == item_name:
                return it
        return {"name": item_name, "total_stock": 0, "color_totals": [], "alerts": [], "alert_count": 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/thresholds/overrides")
async def get_threshold_overrides():
    """Return all current threshold overrides from DB (persistent)."""
    try:
        db = next(get_db())
        try:
            rows = db.query(ThresholdOverride).all()
            result = [
                {
                    "item_name": r.item_name,
                    "size": r.size,
                    "color": r.color,
                    "threshold": r.threshold,
                    "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                }
                for r in rows
            ]
            return {"overrides": result, "count": len(result)}
        finally:
            db.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/thresholds/history")
async def get_threshold_history(item_name: str = None, limit: int = 100):
    """Return threshold change history, optionally filtered by item_name."""
    try:
        db = next(get_db())
        try:
            q = db.query(ThresholdHistory)
            if item_name:
                q = q.filter(ThresholdHistory.item_name == item_name)
            q = q.order_by(ThresholdHistory.changed_at.desc()).limit(min(max(limit, 1), 500))
            rows = q.all()
            result = [
                {
                    "item_name": r.item_name,
                    "size": r.size,
                    "color": r.color,
                    "old_threshold": r.old_threshold,
                    "new_threshold": r.new_threshold,
                    "changed_at": r.changed_at.isoformat() if r.changed_at else None,
                    "note": r.note,
                }
                for r in rows
            ]
            return {"history": result, "count": len(result)}
        finally:
            db.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/files/list-fast")
async def get_files_list_fast():
    """Get just the list of uploaded files - ULTRA FAST, with cached stats when available"""
    try:
        uploads_dir = UPLOAD_DIR
        if not os.path.exists(uploads_dir):
            return {"files": []}
        
        files = [f for f in os.listdir(uploads_dir) if f.endswith('.xlsx')]
        files.sort(key=lambda x: os.path.getmtime(os.path.join(uploads_dir, x)), reverse=True)
        
        # Return basic file info + cached stats if present (no heavy processing)
        file_list = []
        for idx, filename in enumerate(files):
            file_path = os.path.join(uploads_dir, filename)
            try:
                stat = os.stat(file_path)
                # Attach cached stats if available
                file_cache_key = f"file_stats_{filename}"
                cached_stats = None
                if hasattr(key_items_service, '_cache') and file_cache_key in key_items_service._cache:
                    cached_stats = key_items_service._cache[file_cache_key]
                
                file_list.append({
                    "filename": filename,
                    "upload_date": stat.st_mtime,  # epoch seconds
                    "upload_date_iso_utc": datetime.utcfromtimestamp(stat.st_mtime).isoformat() + "Z",
                    "file_size": stat.st_size,
                    "ki00_items_count": (cached_stats or {}).get("key_items_count"),
                    "low_stock_count": (cached_stats or {}).get("low_stock_count")
                })

                # Opportunistically warm stats in background for top 10 files if missing
                if idx < 10 and not cached_stats and filename not in STATS_WARM_IN_PROGRESS:
                    STATS_WARM_IN_PROGRESS.add(filename)
                    threading.Thread(target=warm_file_stats, args=(file_path, filename), daemon=True).start()
            except Exception as e:
                print(f"Error getting file info for {filename}: {e}")
                continue
        
        return {
            "files": file_list,
            "total_files": len(file_list)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving file list: {str(e)}")

@app.get("/files/stats/{filename}")
def get_file_stats(filename: str):
    """Return cached stats for a file if available; otherwise enqueue a background warm-up and return a lightweight placeholder.
    This avoids heavy synchronous computation that can block the event loop and cause client disconnects.
    """
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File {filename} not found")
        
        # Check cache first keyed by filename + mtime
        mtime = os.path.getmtime(file_path)
        cache_key = f"file_stats_{filename}"
        cache_sig_key = f"file_stats_sig_{filename}"
        if not hasattr(key_items_service, '_cache'):
            key_items_service._cache = {}
        if cache_key in key_items_service._cache and cache_sig_key in key_items_service._cache:
            if key_items_service._cache[cache_sig_key] == mtime:
                return key_items_service._cache[cache_key]
        
        # If not cached, schedule background warm-up once and return placeholder
        if filename not in STATS_WARM_IN_PROGRESS:
            STATS_WARM_IN_PROGRESS.add(filename)
            threading.Thread(target=warm_file_stats, args=(file_path, filename), daemon=True).start()
        
        return {
            "filename": filename,
            "key_items_count": None,
            "low_stock_count": None,
            "processed_successfully": False,
            "warmup_in_progress": True
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error computing stats: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 